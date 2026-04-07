"""Spreadsheet IO helpers for CSV and Excel files."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import csv
import posixpath
import xml.etree.ElementTree as ET
import zipfile

from .geo import PolygonBoundary, parse_coordinate, point_in_polygon
from .normalize import guess_lat_lon_columns, is_blank_row, normalize_header


_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XLSX_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass
class SpreadsheetData:
    headers: List[str]
    rows: List[Dict[str, Any]]


@dataclass(frozen=True)
class _XlsxSheetInfo:
    member: str
    headers: List[str]


def read_spreadsheet(path: str) -> SpreadsheetData:
    ext = Path(path).suffix.lower()
    if ext in {".csv"}:
        return _read_csv(path)
    if ext in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ValueError(f"Unsupported file type: {ext}")


def read_spreadsheet_headers(path: str) -> List[str]:
    ext = Path(path).suffix.lower()
    if ext in {".csv"}:
        with open(path, "r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            return [str(item).strip() if item is not None else "" for item in next(reader, [])]
    if ext in {".xlsx", ".xlsm"}:
        try:
            return _read_xlsx_headers_fast(path)
        except Exception:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
            sheet = _select_sheet(workbook)
            return _sheet_headers(sheet)
    raise ValueError(f"Unsupported file type: {ext}")


def read_spreadsheet_preview_points(
    path: str,
    *,
    lat_column: str,
    lon_column: str,
    boundary: PolygonBoundary,
) -> Tuple[List[Tuple[float, float]], int, int, int]:
    ext = Path(path).suffix.lower()
    if ext in {".csv"}:
        return _read_csv_preview_points(path, lat_column=lat_column, lon_column=lon_column, boundary=boundary)
    if ext in {".xlsx", ".xlsm"}:
        try:
            return _read_xlsx_preview_points_fast(path, lat_column=lat_column, lon_column=lon_column, boundary=boundary)
        except Exception:
            return _read_preview_points_from_rows(
                read_spreadsheet(path).rows,
                lat_column=lat_column,
                lon_column=lon_column,
                boundary=boundary,
            )
    raise ValueError(f"Unsupported file type: {ext}")


def write_spreadsheet(path: str, rows: Sequence[Mapping[str, Any]], headers: Sequence[str] | None = None) -> None:
    ext = Path(path).suffix.lower()
    if ext in {".csv"}:
        _write_csv(path, rows, headers=headers)
        return
    if ext in {".xlsx", ".xlsm"}:
        _write_xlsx(path, rows, headers=headers)
        return
    raise ValueError(f"Unsupported file type: {ext}")


def _read_csv(path: str) -> SpreadsheetData:
    with open(path, "r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader if not is_blank_row(row)]
    return SpreadsheetData(headers=headers, rows=rows)


def _write_csv(path: str, rows: Sequence[Mapping[str, Any]], headers: Sequence[str] | None = None) -> None:
    header_list = _resolve_headers(rows, headers)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=header_list)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header) for header in header_list})


def _read_xlsx(path: str) -> SpreadsheetData:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheet = _select_sheet(workbook)
    rows_iter = sheet.iter_rows(values_only=True)
    headers = _sheet_headers(sheet)
    next(rows_iter, None)
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        row_dict: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            if idx < len(row):
                row_dict[header] = row[idx]
        if row_dict and not is_blank_row(row_dict):
            rows.append(row_dict)
    return SpreadsheetData(headers=headers, rows=rows)


def _select_sheet(workbook: Any) -> Any:
    best_sheet = workbook.active
    best_score = (-1, -1)
    for sheet in workbook.worksheets:
        headers = _sheet_headers(sheet)
        nonblank_count = sum(1 for header in headers if header)
        if nonblank_count <= 0:
            continue
        lat_guess, lon_guess = guess_lat_lon_columns(headers)
        crash_like = any(
            token in header.lower()
            for header in headers
            for token in ("crash", "collision", "road", "county", "city", "latitude", "longitude")
        )
        score = (
            (1 if lat_guess and lon_guess else 0) + (1 if crash_like else 0),
            nonblank_count,
        )
        if score > best_score:
            best_sheet = sheet
            best_score = score
    return best_sheet


def _sheet_headers(sheet: Any) -> List[str]:
    row = next(sheet.iter_rows(values_only=True), None)
    if not row:
        return []
    return [str(item).strip() if item is not None else "" for item in row]


def _write_xlsx(path: str, rows: Sequence[Mapping[str, Any]], headers: Sequence[str] | None = None) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl import Workbook

    header_list = _resolve_headers(rows, headers)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(header_list))
    for row in rows:
        sheet.append([row.get(header) for header in header_list])

    if header_list:
        max_row = sheet.max_row
        max_col = len(header_list)
        end_cell = f"{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="Table1", ref=f"A1:{end_cell}")
        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = table_style
        sheet.add_table(table)
        _autosize_columns(sheet, max_col)

    workbook.save(path)


def _autosize_columns(sheet: Any, max_col: int) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, max_col + 1):
        max_length = 0
        for cell in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
            value = cell[0].value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_length:
                max_length = len(text)
        width = min(max(max_length + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _resolve_headers(rows: Sequence[Mapping[str, Any]], headers: Sequence[str] | None) -> List[str]:
    if headers:
        return list(headers)
    header_set = set()
    for row in rows:
        header_set.update(row.keys())
    header_list = sorted(header_set)
    if "kmz_label" in header_list:
        header_list.remove("kmz_label")
        header_list.insert(0, "kmz_label")
    return header_list


def _read_preview_points_from_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    lat_column: str,
    lon_column: str,
    boundary: PolygonBoundary,
) -> Tuple[List[Tuple[float, float]], int, int, int]:
    lat_key = normalize_header(lat_column)
    lon_key = normalize_header(lon_column)
    points: List[Tuple[float, float]] = []
    included = 0
    excluded = 0
    invalid = 0

    for raw_row in rows:
        row = {normalize_header(str(key)): value for key, value in raw_row.items()}
        if is_blank_row(row):
            continue
        lat = parse_coordinate(row.get(lat_key))
        lon = parse_coordinate(row.get(lon_key))
        if lat is None or lon is None:
            invalid += 1
            continue
        if point_in_polygon(lon, lat, boundary):
            points.append((lat, lon))
            included += 1
        else:
            excluded += 1

    return points, included, excluded, invalid


def _read_csv_preview_points(
    path: str,
    *,
    lat_column: str,
    lon_column: str,
    boundary: PolygonBoundary,
) -> Tuple[List[Tuple[float, float]], int, int, int]:
    with open(path, "r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        lat_name = _resolve_requested_header(headers, lat_column)
        lon_name = _resolve_requested_header(headers, lon_column)
        points: List[Tuple[float, float]] = []
        included = 0
        excluded = 0
        invalid = 0

        for row in reader:
            if is_blank_row(row):
                continue
            lat = parse_coordinate(row.get(lat_name))
            lon = parse_coordinate(row.get(lon_name))
            if lat is None or lon is None:
                invalid += 1
                continue
            if point_in_polygon(lon, lat, boundary):
                points.append((lat, lon))
                included += 1
            else:
                excluded += 1

    return points, included, excluded, invalid


def _read_xlsx_headers_fast(path: str) -> List[str]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _xlsx_shared_strings(archive)
        return _xlsx_select_sheet_info(archive, shared_strings).headers


def _read_xlsx_preview_points_fast(
    path: str,
    *,
    lat_column: str,
    lon_column: str,
    boundary: PolygonBoundary,
) -> Tuple[List[Tuple[float, float]], int, int, int]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _xlsx_shared_strings(archive)
        sheet_info = _xlsx_select_sheet_info(archive, shared_strings)
        headers = sheet_info.headers
        lat_name = _resolve_requested_header(headers, lat_column)
        lon_name = _resolve_requested_header(headers, lon_column)
        lat_idx = headers.index(lat_name) if lat_name in headers else -1
        lon_idx = headers.index(lon_name) if lon_name in headers else -1
        points: List[Tuple[float, float]] = []
        included = 0
        excluded = 0
        invalid = 0
        saw_header = False

        with archive.open(sheet_info.member) as handle:
            for _event, elem in ET.iterparse(handle, events=("end",)):
                if elem.tag != _xlsx_tag("row"):
                    continue
                if not saw_header:
                    saw_header = True
                    elem.clear()
                    continue

                has_value = False
                lat_value: Optional[str] = None
                lon_value: Optional[str] = None
                for cell in elem.findall(_xlsx_tag("c")):
                    if _xlsx_cell_has_value(cell):
                        has_value = True
                    col_idx = _xlsx_column_index(cell.attrib.get("r", ""))
                    if col_idx == lat_idx:
                        lat_value = _xlsx_cell_text(cell, shared_strings)
                    elif col_idx == lon_idx:
                        lon_value = _xlsx_cell_text(cell, shared_strings)
                elem.clear()
                if not has_value:
                    continue

                lat = parse_coordinate(lat_value)
                lon = parse_coordinate(lon_value)
                if lat is None or lon is None:
                    invalid += 1
                    continue
                if point_in_polygon(lon, lat, boundary):
                    points.append((lat, lon))
                    included += 1
                else:
                    excluded += 1

    return points, included, excluded, invalid


def _resolve_requested_header(headers: Sequence[str], requested: str) -> str:
    if requested in headers:
        return requested
    normalized = normalize_header(requested)
    for header in headers:
        if normalize_header(header) == normalized:
            return header
    return requested


def _xlsx_select_sheet_info(archive: zipfile.ZipFile, shared_strings: Sequence[str]) -> _XlsxSheetInfo:
    best_info: Optional[_XlsxSheetInfo] = None
    best_score = (-1, -1)
    sheet_members = _xlsx_sheet_members(archive)
    for member in sheet_members:
        headers = _xlsx_sheet_headers(archive, member, shared_strings)
        nonblank_count = sum(1 for header in headers if header)
        if nonblank_count <= 0:
            continue
        lat_guess, lon_guess = guess_lat_lon_columns(headers)
        crash_like = any(
            token in header.lower()
            for header in headers
            for token in ("crash", "collision", "road", "county", "city", "latitude", "longitude")
        )
        score = (
            (1 if lat_guess and lon_guess else 0) + (1 if crash_like else 0),
            nonblank_count,
        )
        if score > best_score:
            best_info = _XlsxSheetInfo(member=member, headers=headers)
            best_score = score
    if best_info is not None:
        return best_info
    if not sheet_members:
        return _XlsxSheetInfo(member="", headers=[])
    return _XlsxSheetInfo(member=sheet_members[0], headers=_xlsx_sheet_headers(archive, sheet_members[0], shared_strings))


def _xlsx_sheet_members(archive: zipfile.ZipFile) -> List[str]:
    with archive.open("xl/workbook.xml") as handle:
        workbook_root = ET.parse(handle).getroot()
    with archive.open("xl/_rels/workbook.xml.rels") as handle:
        rel_root = ET.parse(handle).getroot()

    relationships = {
        rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
        for rel in rel_root.findall(f"{{{_XLSX_PACKAGE_REL_NS}}}Relationship")
    }
    sheets = workbook_root.find(_xlsx_tag("sheets"))
    if sheets is None:
        return []

    members: List[str] = []
    for sheet in sheets.findall(_xlsx_tag("sheet")):
        rel_id = sheet.attrib.get(f"{{{_XLSX_REL_NS}}}id")
        target = relationships.get(rel_id or "")
        if not target:
            continue
        member = _xlsx_archive_member(target)
        if member in archive.namelist():
            members.append(member)
    return members


def _xlsx_sheet_headers(archive: zipfile.ZipFile, member: str, shared_strings: Sequence[str]) -> List[str]:
    if not member:
        return []
    with archive.open(member) as handle:
        for _event, elem in ET.iterparse(handle, events=("end",)):
            if elem.tag != _xlsx_tag("row"):
                continue
            headers: Dict[int, str] = {}
            max_col = -1
            for cell in elem.findall(_xlsx_tag("c")):
                col_idx = _xlsx_column_index(cell.attrib.get("r", ""))
                if col_idx < 0:
                    continue
                headers[col_idx] = _xlsx_cell_text(cell, shared_strings).strip()
                if col_idx > max_col:
                    max_col = col_idx
            elem.clear()
            if max_col < 0:
                return []
            resolved = [""] * (max_col + 1)
            for idx, value in headers.items():
                resolved[idx] = value
            return resolved
    return []


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    try:
        handle = archive.open("xl/sharedStrings.xml")
    except KeyError:
        return []

    strings: List[str] = []
    with handle:
        for _event, elem in ET.iterparse(handle, events=("end",)):
            if elem.tag == _xlsx_tag("si"):
                strings.append(_xlsx_inline_text(elem))
                elem.clear()
    return strings


def _xlsx_cell_text(cell: ET.Element, shared_strings: Sequence[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        inline = cell.find(_xlsx_tag("is"))
        return _xlsx_inline_text(inline) if inline is not None else ""

    value = cell.find(_xlsx_tag("v"))
    if cell_type == "s":
        if value is None or value.text is None:
            return ""
        try:
            index = int(value.text)
        except ValueError:
            return ""
        return shared_strings[index] if 0 <= index < len(shared_strings) else ""
    if value is not None and value.text is not None:
        if cell_type == "b":
            return "TRUE" if value.text == "1" else "FALSE"
        return value.text

    inline = cell.find(_xlsx_tag("is"))
    if inline is not None:
        return _xlsx_inline_text(inline)
    return ""


def _xlsx_cell_has_value(cell: ET.Element) -> bool:
    value = cell.find(_xlsx_tag("v"))
    if value is not None and value.text not in {None, ""}:
        return True
    inline = cell.find(_xlsx_tag("is"))
    if inline is None:
        return False
    return any((text or "") for text in inline.itertext())


def _xlsx_inline_text(element: Optional[ET.Element]) -> str:
    if element is None:
        return ""
    return "".join(text or "" for text in element.itertext())


def _xlsx_column_index(cell_ref: str) -> int:
    letters: List[str] = []
    for char in cell_ref:
        if char.isalpha():
            letters.append(char.upper())
            continue
        break
    if not letters:
        return -1

    value = 0
    for char in letters:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value - 1


def _xlsx_archive_member(target: str) -> str:
    normalized = target.replace("\\", "/")
    if normalized.startswith("/"):
        return normalized.lstrip("/")
    if normalized.startswith("xl/"):
        return posixpath.normpath(normalized)
    return posixpath.normpath(posixpath.join("xl", normalized))


def _xlsx_tag(name: str) -> str:
    return f"{{{_XLSX_MAIN_NS}}}{name}"
