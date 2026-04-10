"""Tests for the shared normalize and services modules."""
from __future__ import annotations

from pathlib import Path
import zipfile

from crash_data_refiner.coordinate_recovery import CoordinateReviewDecision
from crash_data_refiner.geo import load_kmz_polygon
from crash_data_refiner.normalize import normalize_header, guess_lat_lon_columns
from crash_data_refiner.spreadsheets import (
    read_spreadsheet,
    read_spreadsheet_headers,
    read_spreadsheet_preview_points,
    write_spreadsheet,
)
from crash_data_refiner.services import (
    coordinate_review_output_path,
    detect_label_order,
    refined_output_path,
    invalid_output_path,
    kmz_output_path,
    relabel_refined_outputs,
    rejected_review_output_path,
    build_output_headers,
    order_and_number_rows,
    resolve_label_order,
    run_refinement_pipeline,
)


_UNIT_KML = """<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <Placemark>
      <Polygon>
        <outerBoundaryIs>
          <LinearRing>
            <coordinates>
              -1,0 1,0 1,2 -1,2 -1,0
            </coordinates>
          </LinearRing>
        </outerBoundaryIs>
      </Polygon>
    </Placemark>
  </Document>
</kml>
"""


# ---------------------------------------------------------------------------
# normalize_header
# ---------------------------------------------------------------------------


def test_normalize_header_basic() -> None:
    assert normalize_header("Crash ID") == "crash_id"
    assert normalize_header("Crash Date") == "crash_date"
    assert normalize_header("  Hit and Run  ") == "hit_and_run"


def test_normalize_header_special_chars() -> None:
    assert normalize_header("Lat/Long") == "lat_long"
    assert normalize_header("__foo__") == "foo"


# ---------------------------------------------------------------------------
# guess_lat_lon_columns
# ---------------------------------------------------------------------------


def test_guess_lat_lon_exact_match() -> None:
    lat, lon = guess_lat_lon_columns(["Crash ID", "Latitude", "Longitude", "City"])
    assert lat == "Latitude"
    assert lon == "Longitude"


def test_guess_lat_lon_short_names() -> None:
    lat, lon = guess_lat_lon_columns(["lat", "lon"])
    assert lat == "lat"
    assert lon == "lon"


def test_guess_lat_lon_no_match() -> None:
    lat, lon = guess_lat_lon_columns(["Crash ID", "City"])
    assert lat is None
    assert lon is None


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------


def test_refined_output_path(tmp_path: Path) -> None:
    p = refined_output_path(tmp_path, "crashes.csv")
    assert p == tmp_path / "crashes_refined.csv"


def test_refined_output_path_xlsx(tmp_path: Path) -> None:
    p = refined_output_path(tmp_path, "data.xlsx")
    assert p == tmp_path / "data_refined.xlsx"


def test_invalid_output_path(tmp_path: Path) -> None:
    base = tmp_path / "crashes_refined.csv"
    p = invalid_output_path(base)
    assert p.name == "Crashes Without Valid Lat-Long Data.csv"


def test_coordinate_review_output_path(tmp_path: Path) -> None:
    base = tmp_path / "crashes_refined.csv"
    p = coordinate_review_output_path(base)
    assert p.name == "crashes_Coordinate Recovery Review.csv"


def test_rejected_review_output_path(tmp_path: Path) -> None:
    base = tmp_path / "crashes_refined.csv"
    p = rejected_review_output_path(base)
    assert p.name == "crashes_Rejected Coordinate Review.csv"


def test_kmz_output_path_strips_refined(tmp_path: Path) -> None:
    base = tmp_path / "crashes_refined.csv"
    p = kmz_output_path(base)
    assert p.name == "crashes_Crash Data.kmz"
# ---------------------------------------------------------------------------
# Spreadsheet sheet selection
# ---------------------------------------------------------------------------


def test_read_spreadsheet_prefers_data_sheet_over_blank_about_sheet(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "multi_sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "About"
    workbook.create_sheet("Target Collision Records")
    sheet = workbook["Target Collision Records"]
    sheet.append(["Crash ID", "Latitude", "Longitude"])
    sheet.append(["1", "40.0", "-86.0"])
    workbook.save(path)

    headers = read_spreadsheet_headers(str(path))
    data = read_spreadsheet(str(path))

    assert headers == ["Crash ID", "Latitude", "Longitude"]
    assert data.headers == headers
    assert data.rows == [{"Crash ID": "1", "Latitude": "40.0", "Longitude": "-86.0"}]


def test_read_spreadsheet_preview_points_filters_xlsx_rows(tmp_path: Path) -> None:
    from openpyxl import Workbook

    data_path = tmp_path / "preview.xlsx"
    workbook = Workbook()
    workbook.active.title = "About"
    workbook.active.append(["Notes"])
    sheet = workbook.create_sheet("Crash Data")
    sheet.append(["Crash ID", "Latitude", "Longitude"])
    sheet.append(["1", "1.0", "0.0"])
    sheet.append(["2", "3.0", "0.0"])
    sheet.append(["3", "", ""])
    workbook.save(data_path)

    kmz_path = tmp_path / "boundary.kmz"
    _write_test_kmz(kmz_path)
    boundary = load_kmz_polygon(str(kmz_path))

    points, included, excluded, invalid = read_spreadsheet_preview_points(
        str(data_path),
        lat_column="Latitude",
        lon_column="Longitude",
        boundary=boundary,
    )

    assert points == [(1.0, 0.0)]
    assert included == 1
    assert excluded == 1
    assert invalid == 1


# ---------------------------------------------------------------------------
# build_output_headers
# ---------------------------------------------------------------------------


def test_build_output_headers_kmz_label_first() -> None:
    rows = [{"kmz_label": 1, "b": 2, "a": 3}]
    headers = build_output_headers(rows)
    assert headers[0] == "kmz_label"
    assert sorted(headers[1:]) == sorted(["a", "b"])


def test_build_output_headers_no_kmz_label() -> None:
    rows = [{"b": 1, "a": 2}]
    headers = build_output_headers(rows)
    assert headers == ["a", "b"]


# ---------------------------------------------------------------------------
# order_and_number_rows
# ---------------------------------------------------------------------------


def test_order_south_to_north() -> None:
    rows = [
        {"lat": "40.0", "lon": "-100.0"},
        {"lat": "30.0", "lon": "-100.0"},
        {"lat": "35.0", "lon": "-100.0"},
    ]
    ordered = order_and_number_rows(rows, lat_column="lat", lon_column="lon", label_order="south_to_north")
    lats = [float(r["lat"]) for r in ordered]
    assert lats == sorted(lats)
    assert [r["kmz_label"] for r in ordered] == [1, 2, 3]


def test_order_west_to_east() -> None:
    rows = [
        {"lat": "35.0", "lon": "-90.0"},
        {"lat": "35.0", "lon": "-110.0"},
        {"lat": "35.0", "lon": "-100.0"},
    ]
    ordered = order_and_number_rows(rows, lat_column="lat", lon_column="lon", label_order="west_to_east")
    lons = [float(r["lon"]) for r in ordered]
    assert lons == sorted(lons)


def test_detect_label_order_prefers_west_to_east_when_longitude_span_is_larger() -> None:
    rows = [
        {"lat": "40.0", "lon": "-110.0"},
        {"lat": "40.3", "lon": "-100.0"},
        {"lat": "40.1", "lon": "-90.0"},
    ]

    assert detect_label_order(rows, lat_column="lat", lon_column="lon") == "west_to_east"


def test_detect_label_order_prefers_south_to_north_when_latitude_span_is_larger() -> None:
    rows = [
        {"lat": "30.0", "lon": "-100.0"},
        {"lat": "40.0", "lon": "-100.5"},
        {"lat": "50.0", "lon": "-100.2"},
    ]

    assert detect_label_order(rows, lat_column="lat", lon_column="lon") == "south_to_north"


def test_resolve_label_order_defaults_ties_to_west_to_east() -> None:
    rows = [
        {"lat": "30.0", "lon": "-100.0"},
        {"lat": "40.0", "lon": "-90.0"},
    ]

    assert resolve_label_order(rows, lat_column="lat", lon_column="lon", label_order="auto") == "west_to_east"


# ---------------------------------------------------------------------------
# run_refinement_pipeline integration
# ---------------------------------------------------------------------------


def _write_test_kmz(path: Path) -> None:
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("doc.kml", _UNIT_KML)


def test_run_refinement_pipeline(tmp_path: Path) -> None:
    import csv

    # Write a crash CSV with one point inside, one outside, and one missing row
    # that should be auto-recovered from the exact same location signature.
    data_file = tmp_path / "crashes.csv"
    with data_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Crash ID", "Lat", "Lon", "Roadway Number", "Intersecting Road", "City", "County"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "Crash ID": "1",
                "Lat": "1.0",
                "Lon": "0.0",
                "Roadway Number": "SR1",
                "Intersecting Road": "Main",
                "City": "Testville",
                "County": "Example",
            }
        )  # inside
        writer.writerow({"Crash ID": "2", "Lat": "5.0", "Lon": "0.0"})  # outside
        writer.writerow(
            {
                "Crash ID": "3",
                "Lat": "",
                "Lon": "",
                "Roadway Number": "SR1",
                "Intersecting Road": "Main",
                "City": "Testville",
                "County": "Example",
            }
        )  # auto-recovered inside

    kmz_file = tmp_path / "boundary.kmz"
    _write_test_kmz(kmz_file)

    run_dir = tmp_path / "run_001"
    result = run_refinement_pipeline(
        data_path=data_file,
        kmz_path=kmz_file,
        run_dir=run_dir,
        lat_column="Lat",
        lon_column="Lon",
    )

    assert result.boundary_report.included_rows == 2
    assert result.boundary_report.excluded_rows == 1
    assert result.output_path.exists()
    assert result.coordinate_review_path.exists()
    assert result.kmz_path.exists()
    assert result.kmz_count == 2
    assert result.requested_label_order == "auto"
    assert result.resolved_label_order == "west_to_east"
    assert result.recovery_report.missing_rows == 1
    assert result.recovery_report.recovered_rows == 1
    assert result.recovery_report.review_rows == 0

    refined = read_spreadsheet(str(result.output_path))
    recovered_row = next(row for row in refined.rows if row["crash_id"] == "3")
    assert float(recovered_row["lat"]) == 1.0
    assert float(recovered_row["lon"]) == 0.0
    assert recovered_row["coordinate_source"] == "recovered"


def test_relabel_refined_outputs_rewrites_refined_file(tmp_path: Path) -> None:
    refined_path = tmp_path / "crashes_refined.csv"
    write_spreadsheet(
        str(refined_path),
        [
            {"crash_id": "3", "lat": "40.0", "lon": "-86.0", "kmz_label": 99},
            {"crash_id": "1", "lat": "30.0", "lon": "-86.1", "kmz_label": 98},
            {"crash_id": "2", "lat": "35.0", "lon": "-86.2", "kmz_label": 97},
        ],
    )
    kmz_path = tmp_path / "crashes_Crash Data.kmz"

    result = relabel_refined_outputs(
        refined_path=refined_path,
        kmz_path=kmz_path,
        lat_column="lat",
        lon_column="lon",
        label_order="south_to_north",
    )

    assert result.requested_label_order == "south_to_north"
    assert result.resolved_label_order == "south_to_north"
    assert result.kmz_path.exists()
    assert result.removed_outputs == []

    refined = read_spreadsheet(str(refined_path))
    assert [row["crash_id"] for row in refined.rows] == ["1", "2", "3"]
    assert [row["kmz_label"] for row in refined.rows] == ["1", "2", "3"]


def test_run_refinement_pipeline_applies_coordinate_review_workbook(tmp_path: Path) -> None:
    import csv

    data_file = tmp_path / "crashes.csv"
    with data_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Crash ID", "Lat", "Lon", "Roadway Number", "Intersecting Road", "City", "County"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "Crash ID": "1",
                "Lat": "1.0",
                "Lon": "0.0",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )  # inside boundary
        writer.writerow(
            {
                "Crash ID": "2",
                "Lat": "5.0",
                "Lon": "0.0",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )  # outside boundary
        writer.writerow(
            {
                "Crash ID": "3",
                "Lat": "",
                "Lon": "",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )  # unresolved until review workbook is applied

    review_file = tmp_path / "coordinate_review.csv"
    with review_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "coordinate_recovery_group",
                "approve_for_group",
                "approved_latitude",
                "approved_longitude",
                "review_notes",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "coordinate_recovery_group": "SR2|OAK|TESTVILLE|EXAMPLE",
                "approve_for_group": "yes",
                "approved_latitude": "1.0",
                "approved_longitude": "0.0",
                "review_notes": "Approved grouped location.",
            }
        )

    kmz_file = tmp_path / "boundary.kmz"
    _write_test_kmz(kmz_file)

    run_dir = tmp_path / "run_002"
    result = run_refinement_pipeline(
        data_path=data_file,
        kmz_path=kmz_file,
        run_dir=run_dir,
        lat_column="Lat",
        lon_column="Lon",
        coordinate_review_path=review_file,
    )

    assert result.boundary_report.included_rows == 2
    assert result.boundary_report.excluded_rows == 1
    assert result.boundary_report.invalid_rows == 0
    assert result.recovery_report.missing_rows == 1
    assert result.recovery_report.recovered_rows == 1
    assert result.recovery_report.approved_rows == 1
    assert result.recovery_report.review_rows == 0

    refined = read_spreadsheet(str(result.output_path))
    recovered_row = next(row for row in refined.rows if row["crash_id"] == "3")
    assert float(recovered_row["lat"]) == 1.0
    assert float(recovered_row["lon"]) == 0.0
    assert recovered_row["coordinate_source"] == "review_approved"
    assert recovered_row["coordinate_recovery_status"] == "review_applied"


def test_run_refinement_pipeline_applies_browser_review_decisions(tmp_path: Path) -> None:
    import csv

    data_file = tmp_path / "crashes.csv"
    with data_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Crash ID", "Lat", "Lon", "Roadway Number", "Intersecting Road", "City", "County"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "Crash ID": "1",
                "Lat": "",
                "Lon": "",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )

    kmz_file = tmp_path / "boundary.kmz"
    _write_test_kmz(kmz_file)

    run_dir = tmp_path / "run_003"
    result = run_refinement_pipeline(
        data_path=data_file,
        kmz_path=kmz_file,
        run_dir=run_dir,
        lat_column="Lat",
        lon_column="Lon",
        review_decisions={
            "SR2|OAK|TESTVILLE|EXAMPLE": CoordinateReviewDecision(
                group_key="SR2|OAK|TESTVILLE|EXAMPLE",
                latitude=1.0,
                longitude=0.0,
                note="Applied from browser review queue.",
            )
        },
    )

    assert result.recovery_report.approved_rows == 1
    assert result.recovery_report.recovered_rows == 1
    assert result.boundary_report.included_rows == 1
    refined = read_spreadsheet(str(result.output_path))
    recovered_row = refined.rows[0]
    assert recovered_row["coordinate_source"] == "review_approved"


def test_run_refinement_pipeline_applies_row_level_browser_review_decisions_and_rejections(
    tmp_path: Path,
) -> None:
    import csv

    data_file = tmp_path / "crashes.csv"
    with data_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Crash ID", "Lat", "Lon", "Roadway Number", "Intersecting Road", "City", "County"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "Crash ID": "1",
                "Lat": "1.0000",
                "Lon": "0.0000",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )
        writer.writerow(
            {
                "Crash ID": "2",
                "Lat": "1.0006",
                "Lon": "0.0000",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )
        writer.writerow(
            {
                "Crash ID": "3",
                "Lat": "",
                "Lon": "",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )
        writer.writerow(
            {
                "Crash ID": "4",
                "Lat": "",
                "Lon": "",
                "Roadway Number": "SR2",
                "Intersecting Road": "Oak",
                "City": "Testville",
                "County": "Example",
            }
        )

    kmz_file = tmp_path / "boundary.kmz"
    _write_test_kmz(kmz_file)

    run_dir = tmp_path / "run_004"
    result = run_refinement_pipeline(
        data_path=data_file,
        kmz_path=kmz_file,
        run_dir=run_dir,
        lat_column="Lat",
        lon_column="Lon",
        review_decisions={
            "3__row4": CoordinateReviewDecision(
                group_key="3__row4",
                latitude=1.0003,
                longitude=0.0,
                action="apply",
                note="Confirmed manual map placement.",
            ),
            "4__row5": CoordinateReviewDecision(
                group_key="4__row5",
                action="reject",
                note="Excluded during browser review workbench session.",
            ),
        },
    )

    assert result.recovery_report.missing_rows == 2
    assert result.recovery_report.approved_rows == 1
    assert result.recovery_report.rejected_rows == 1
    assert result.recovery_report.review_rows == 0
    assert result.boundary_report.included_rows == 3
    assert result.invalid_rows == []
    assert len(result.rejected_review_rows) == 1
    assert result.rejected_review_path.exists()

    refined = read_spreadsheet(str(result.output_path))
    refined_ids = {row["crash_id"] for row in refined.rows}
    assert refined_ids == {"1", "2", "3"}
    applied_row = next(row for row in refined.rows if row["crash_id"] == "3")
    assert applied_row["coordinate_source"] == "review_approved"
    assert applied_row["coordinate_recovery_status"] == "review_applied"

    rejected = read_spreadsheet(str(result.rejected_review_path))
    assert len(rejected.rows) == 1
    assert rejected.rows[0]["crash_id"] == "4"
    assert rejected.rows[0]["coordinate_recovery_status"] == "review_rejected"
    assert rejected.rows[0]["coordinate_source"] == "review_rejected"
